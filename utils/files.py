# -*- coding: utf-8 -*-
"""
| Author: Alexandre CARRE (alexandre.carre@gustaveroussy.fr)
| Created on: Nov 23, 2020
"""
import glob
import logging
import os
from collections import OrderedDict
from typing import Tuple, List, Dict, Optional

import SimpleITK as sitk
import numpy as np
import pandas as pd
import torch
from monai.data import write_nifti
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def check_isdir(input_dir: str) -> str:
    """
    Check if a directory exist.

    :param input_dir: string of the path of the input directory.
    :return: string if exist, else raise NotADirectoryError.
    """
    if os.path.isdir(input_dir):
        return input_dir
    else:
        raise NotADirectoryError(input_dir)


def check_exist(input_file: str) -> str:
    """
    Check if a file exist

    :param input_file: string of the path of the input file.
    :return: string if exist, else raise FileNotFoundError.
    """
    if os.path.exists(input_file):
        return input_file
    else:
        raise FileNotFoundError(input_file)


def check_is_nii_exist(input_file_path: str) -> str:
    """
    Check if a directory exist.

    :param input_file_path: string of the path of the nii or nii.gz.
    :return: string if exist, else raise Error.
    """
    if not os.path.exists(input_file_path):
        raise FileNotFoundError(f"{input_file_path} was not found, check if it's a valid file path")

    pth, fnm, ext = split_filename(input_file_path)
    if ext not in [".nii", ".nii.gz"]:
        raise FileExistsError(f"extension of {input_file_path} need to be '.nii' or '.nii.gz'")
    return input_file_path


def safe_file_name(file_name: str) -> str:
    """
    Remove any potentially dangerous or confusing characters from
    the file name by mapping them to reasonable substitutes.

    :param file_name: name of the file.
    :return: name of the file corrected.
    """
    underscores = r"""+`~!?@#$%^&*(){}[]/=\|<>,.":' """
    safe_name = ""
    for c in file_name:
        if c in underscores:
            c = "_"
        safe_name += c
    return safe_name


def split_filename(file_name: str) -> Tuple[str, str, str]:
    """
    Split file_name into folder path name, basename, and extension name.

    :param file_name: full path
    :return: path name, basename, extension name
    """
    pth = os.path.dirname(file_name)
    f_name = os.path.basename(file_name)

    ext = None
    for special_ext in ['.nii.gz']:
        ext_len = len(special_ext)
        if f_name[-ext_len:].lower() == special_ext:
            ext = f_name[-ext_len:]
            f_name = f_name[:-ext_len] if len(f_name) > ext_len else ''
            break
    if not ext:
        f_name, ext = os.path.splitext(f_name)
    return pth, f_name, ext


def load_nifty_volume_as_array(input_path_file: str) -> Tuple[np.ndarray, Tuple[Tuple, Tuple, Tuple]]:
    """
    Load nifty image into numpy array [z,y,x] axis order.
    The output array shape is like [Depth, Height, Width].

    :param input_path_file: input path file, should be '*.nii' or '*.nii.gz'
    :return: a numpy data array, (with header)
    """
    img = sitk.ReadImage(input_path_file)
    data = sitk.GetArrayFromImage(img)

    origin, spacing, direction = img.GetOrigin(), img.GetSpacing(), img.GetDirection()
    return data, (origin, spacing, direction)


def save_to_nii(im: np.ndarray, header: (tuple, tuple, tuple), output_dir: str, filename: str, mode: str = "image",
                gzip: bool = True) -> None:
    """
    Save numpy array to nii.gz format to submit.

    :param im: array numpy
    :param header: header metadata (origin, spacing, direction).
    :param output_dir: Output directory.
    :param filename: Filename of the output file.
    :param mode: save as 'image' or 'label'
    :param gzip: zip nii (ie, nii.gz)
    """
    origin, spacing, direction = header
    if mode == "seg":
        img = sitk.GetImageFromArray(im.astype(np.uint8))
    else:
        img = sitk.GetImageFromArray(im.astype(np.float32))
    img.SetOrigin(origin), img.SetSpacing(spacing), img.SetDirection(direction)

    if gzip:
        ext = ".nii.gz"
    else:
        ext = ".nii"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)
    sitk.WriteImage(img, os.path.join(output_dir, filename) + ext)


def create_database(input_dir: str, filter_keyword: str = None,
                    filtering_patient: List[str] = None,
                    required_modality: List[str] = ("t1", "t1ce", "flair", "t2"),
                    keep_going: bool = True) -> Dict:
    """
    Create database patient as a dict. {patient_ID: {t1ce: path, t1: path, flair: path}}

    :param required_modality: required modality. Split on last suffix '_' on file base name
                              (default : ["t1", "t1ce", "flair", "t2"])
    :param input_dir: input directory
    :param filter_keyword: keyword to filter files in input_dir
    :param keep_going: Set to True to not take into account the patient folder if there is missing modalities else
                       :raise ValueError
    :return: dict of patient database
    """
    patient_dict = OrderedDict()
    for directory in sorted(os.listdir(input_dir)):
        if os.path.isdir(os.path.join(input_dir, directory)):
            native_image, patient_modalities = {}, []
            for file in glob.glob(os.path.join(input_dir, directory, "**", "*.nii*"), recursive=True):
                _, fnm, _ = split_filename(file)
                if filter_keyword is not None:
                    if filter_keyword not in fnm:
                        continue
                if filtering_patient is not None:
                    if any([x in file for x in filtering_patient]):
                        continue
                expected_modality = fnm.split('_')[-1]
                if expected_modality in required_modality:
                    patient_modalities.append(expected_modality)
                    native_image[expected_modality] = os.path.join(input_dir, directory, file)
            if not all(x in patient_modalities for x in required_modality):
                if not keep_going:
                    raise ValueError(f"The patient {directory} has missing modalities \n"
                                     f"find: {patient_modalities}")
                logger.warning(f"Not take into account the patient {directory} because there is missing modalities"
                               f" find: {patient_modalities}")
            else:
                patient_dict[directory] = native_image

    # reorder by required modality order
    patient_dict = OrderedDict({k: {mod: v[mod] for mod in required_modality} for k, v in patient_dict.items()})

    return patient_dict


def create_database_test_docker(input_dir: str, filter_keyword: str = None,
                                filtering_patient: List[str] = None,
                                required_modality: List[str] = ("t1", "t1ce", "flair", "t2"),
                                ) -> Dict:
    """
    Create database patient as a dict. {patient_ID: {t1ce: path, t1: path, flair: path}}

    :param required_modality: required modality. Split on last suffix '_' on file base name
                              (default : ["t1", "t1ce", "flair", "t2"])
    :param input_dir: input directory
    :param filter_keyword: keyword to filter files in input_dir
    :param keep_going: Set to True to not take into account the patient folder if there is missing modalities else
                       :raise ValueError
    :return: dict of patient database
    """
    patient_dict = OrderedDict()
    native_image, patient_modalities = {}, []
    name = ""
    for file in glob.glob(os.path.join(input_dir, "**", "*.nii*"), recursive=True):
        _, fnm, _ = split_filename(file)
        if filter_keyword is not None:
            if filter_keyword not in fnm:
                continue
        if filtering_patient is not None:
            if any([x in file for x in filtering_patient]):
                continue
        expected_modality = fnm.split('_')[-1]
        name = "_".join(fnm.split('_')[:-1])
        if expected_modality in required_modality:
            patient_modalities.append(expected_modality)
            native_image[expected_modality] = os.path.join(input_dir, file)

    patient_dict[name] = native_image

    # reorder by required modality order
    patient_dict = OrderedDict({k: {mod: v[mod] for mod in required_modality} for k, v in patient_dict.items()})

    return patient_dict


def segmentation_save(batch_data: Dict, outputs: np.ndarray, output_path: str, suffix: Optional[str] = None) -> None:
    """
    Pad back to the original shape before apply monai.transforms.Compose

    :param batch_data: batch data from the data loader
    :param outputs: outputs pred tensor
    :param output_path: path where to save seg
    :param suffix: suffix to add to filename
    """
    if not os.path.exists(output_path):
        os.makedirs(output_path, exist_ok=True)

    filename = os.path.join(output_path, batch_data["patient_id"][
        0] + '_' + suffix + ".nii.gz") if suffix else os.path.join(output_path,
                                                                   batch_data["patient_id"][
                                                                       0] + ".nii.gz")
    write_nifti(data=outputs,
                file_name=filename,
                affine=np.squeeze(batch_data["img_meta_dict"]["affine"]),
                target_affine=np.squeeze(batch_data["img_meta_dict"]["original_affine"]),
                resample=False, output_dtype=np.uint8)


def save_checkpoint(filepath: str, epoch: int, **kwargs) -> None:
    state = {
        'epoch': epoch,
    }
    state.update(kwargs)
    torch.save(state, filepath)


def load_checkpoint(filepath: str) -> Dict:
    state = torch.load(filepath)
    return state


def append_df_to_excel(filename: str, df: pd.DataFrame, sheet_name: str = 'Sheet1', start_row: Optional = None,
                       truncate_sheet: bool = False, **to_excel_kwargs) -> None:
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    .. seealso::
        `<https://stackoverflow.com/questions/47737220/append-dataframe-to-excel-with-pandas>`_

    :param filename: File path or existing ExcelWriter
        (Example: '/path/to/file.xlsx')
    :param df: dataframe to save to workbook
    :param sheet_name : Name of sheet which will contain DataFrame.
        (default: 'Sheet1')
    :param start_row: upper left cell row to dump data frame.
        Per default (startrow=None) calculate the last row
        in the existing DF and write to the next row...
    :param truncate_sheet: truncate (remove and recreate) [sheet_name]
        before writing DataFrame to Excel file
    :param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
        [can be dictionary]

    :return: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if start_row is None and sheet_name in writer.book.sheetnames:
            start_row = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if start_row is None:
        start_row = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=start_row, **to_excel_kwargs)
    df.to_excel(writer, sheet_name, startrow=start_row, **to_excel_kwargs)

    # save the workbook
    writer.save()
